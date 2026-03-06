import os
import re
import sys
import time
import base64
import json
from datetime import datetime

print("Bot module loading...", flush=True)

import httpx
from dotenv import load_dotenv
from telegram import Update, Bot, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters

load_dotenv()

TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
ALLOWED_USER_ID = int(os.environ["TELEGRAM_USER_ID"])
GITHUB_TOKEN = os.environ["GITHUB_TOKEN"]
GITHUB_REPO = os.environ.get("GITHUB_REPO", "amijkko/personal-goals")
OPENAI_API_KEY = os.environ["OPENAI_API_KEY"]
BACKLOG_FILE = "backlog.md"

GITHUB_BASE = f"https://api.github.com/repos/{GITHUB_REPO}/contents"
GITHUB_API = f"{GITHUB_BASE}/{BACKLOG_FILE}"
GH_HEADERS = {
    "Authorization": f"Bearer {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json",
}

OPENAI_HEADERS = {
    "Authorization": f"Bearer {OPENAI_API_KEY}",
    "Content-Type": "application/json",
}

ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
ANTHROPIC_HEADERS = {
    "x-api-key": ANTHROPIC_API_KEY,
    "anthropic-version": "2023-06-01",
    "Content-Type": "application/json",
}
CLAUDE_MODEL = "claude-sonnet-4-20250514"

USE_CLAUDE = True


def claude_complete(prompt: str, max_tokens: int = 500, temperature: float = 0.2) -> str:
    """Call LLM API (Claude or GPT fallback) and return text response."""
    if USE_CLAUDE:
        r = httpx.post(
            "https://api.anthropic.com/v1/messages",
            headers=ANTHROPIC_HEADERS,
            json={
                "model": CLAUDE_MODEL,
                "max_tokens": max_tokens,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": temperature,
            },
            timeout=30,
        )
        if r.status_code != 200:
            print(f"Claude API error {r.status_code}: {r.text}", flush=True)
        r.raise_for_status()
        return r.json()["content"][0]["text"].strip()
    else:
        r = httpx.post(
            "https://api.openai.com/v1/chat/completions",
            headers=OPENAI_HEADERS,
            json={
                "model": "gpt-4o-mini",
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": max_tokens,
                "temperature": temperature,
            },
            timeout=30,
        )
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"].strip()


# --- Priority detection ---

URGENT_KEYWORDS = r"(?:срочно|важно|асап|asap|urgent|горит|критично)"
LOW_KEYWORDS = r"(?:не\s*важно|неважно|когда-нибудь|потом|low|при\s*случае|без\s*спешки)"

SECTION_MAP = {
    "urgent": "## Urgent (this week)",
    "normal": "## Important (next 2-3 weeks)",
    "low": "## Someday (no deadline)",
}

PRIORITY_LABELS = {
    "urgent": "Urgent",
    "normal": "Important",
    "low": "Someday",
}

# --- Projects for KB ---

PROJECTS = {
    "custody": {"name": "Custody", "path": "kb/custody", "crm": "crm.md"},
    "sber": {"name": "Сбер Стейблкоин", "path": "kb/sber", "crm": "crm-sber.md"},
    "reksoft": {"name": "Reksoft Consulting", "path": "kb/reksoft", "crm": "crm.md"},
    "blind-bets": {"name": "Blind Bets", "path": "kb/blind-bets", "crm": "crm-blind-bets.md"},
}

PROJECT_KEYWORDS = {
    "custody": ["custody", "кастоди", "кастодиан", "mpc", "инвестор", "фандрайзинг", "crm"],
    "sber": ["сбер", "стейблкоин", "stablecoin", "гаймаков", "sber"],
    "reksoft": ["reksoft", "рексофт", "консалтинг", "скорочкин", "артём", "артем"],
    "blind-bets": ["blind", "bets", "ставки", "денис", "autoforge"],
}


WEEKDAYS_RU = {
    "понедельник": 0, "вторник": 1, "среду": 2, "среда": 2,
    "четверг": 3, "пятницу": 4, "пятница": 4,
    "субботу": 5, "суббота": 5, "воскресенье": 6,
}

MONTHS_RU = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4,
    "мая": 5, "июня": 6, "июля": 7, "августа": 8,
    "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
}

DAY_NAMES_SHORT = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]


def parse_date(text: str) -> tuple[str | None, str]:
    """Extract date from Russian text. Returns (date_label, cleaned_text)."""
    from datetime import timedelta
    today = datetime.now().date()
    lower = text.lower()

    # "сегодня"
    m = re.search(r"\bсегодня\b", lower)
    if m:
        d = today
        clean = text[:m.start()] + text[m.end():]
        label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
        return label, clean.strip(" ,:-—")

    # "завтра"
    m = re.search(r"\bзавтра\b", lower)
    if m:
        d = today + timedelta(days=1)
        clean = text[:m.start()] + text[m.end():]
        label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
        return label, clean.strip(" ,:-—")

    # "послезавтра"
    m = re.search(r"\bпослезавтра\b", lower)
    if m:
        d = today + timedelta(days=2)
        clean = text[:m.start()] + text[m.end():]
        label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
        return label, clean.strip(" ,:-—")

    # "через N дней/дня"
    m = re.search(r"\bчерез\s+(\d+)\s+(?:день|дня|дней)\b", lower)
    if m:
        d = today + timedelta(days=int(m.group(1)))
        clean = text[:m.start()] + text[m.end():]
        label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
        return label, clean.strip(" ,:-—")

    # "в понедельник" / "на среду" / "во вторник" (next occurrence)
    for day_name, day_num in WEEKDAYS_RU.items():
        pattern = rf"\b(?:в|на|во)\s+(?:след(?:ующ[а-я]*)?\s+)?{day_name}\b"
        m = re.search(pattern, lower)
        if m:
            is_next = "след" in m.group(0)
            days_ahead = (day_num - today.weekday()) % 7
            if days_ahead == 0:
                days_ahead = 7
            if is_next and days_ahead < 7:
                days_ahead += 7
            d = today + timedelta(days=days_ahead)
            clean = text[:m.start()] + text[m.end():]
            label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
            return label, clean.strip(" ,:-—")

    # "на следующей неделе" (Monday)
    m = re.search(r"\bна\s+след(?:ующ[а-я]*)?\s+недел[а-яё]+\b", lower)
    if m:
        days_ahead = (0 - today.weekday()) % 7
        if days_ahead == 0:
            days_ahead = 7
        d = today + timedelta(days=days_ahead)
        clean = text[:m.start()] + text[m.end():]
        label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
        return label, clean.strip(" ,:-—")

    # "N марта" / "N мая" etc
    for month_name, month_num in MONTHS_RU.items():
        pattern = rf"\b(\d{{1,2}})\s+{month_name}\b"
        m = re.search(pattern, lower)
        if m:
            day = int(m.group(1))
            year = today.year
            try:
                from datetime import date
                d = date(year, month_num, day)
                if d < today:
                    d = date(year + 1, month_num, day)
            except ValueError:
                continue
            clean = text[:m.start()] + text[m.end():]
            label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
            return label, clean.strip(" ,:-—")

    # "в понедельник" without preposition — just weekday name
    for day_name, day_num in WEEKDAYS_RU.items():
        m = re.search(rf"\b{day_name}\b", lower)
        if m:
            days_ahead = (day_num - today.weekday()) % 7
            if days_ahead == 0:
                days_ahead = 7
            d = today + timedelta(days=days_ahead)
            clean = text[:m.start()] + text[m.end():]
            label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
            return label, clean.strip(" ,:-—")

    return None, text


CRM_FILE = "crm.md"


def parse_crm(content: str) -> list[dict]:
    """Parse crm.md into structured contacts."""
    contacts = []
    current = None
    for line in content.split("\n"):
        line_s = line.strip()
        if line_s.startswith("### "):
            if current:
                contacts.append(current)
            # "### Name — Company"
            parts = line_s[4:].split("—", 1)
            current = {
                "name": parts[0].strip(),
                "company": parts[1].strip() if len(parts) > 1 else "",
                "contact": "",
                "warmth": "",
                "confidence": "",
                "next_action": "",
                "why": "",
                "comment": "",
                "ping_date": None,
                "raw_lines": [line],
            }
        elif current:
            current["raw_lines"].append(line)
            if "**Контакт:**" in line_s:
                current["contact"] = line_s.split("**Контакт:**")[1].strip()
            elif "**Теплота:**" in line_s:
                m = re.search(r"\*\*Теплота:\*\*\s*([^|]+)", line_s)
                if m:
                    current["warmth"] = m.group(1).strip()
                m2 = re.search(r"\*\*Уверенность:\*\*\s*(\d+)", line_s)
                if m2:
                    current["confidence"] = m2.group(1) + "%"
            elif "**Следующее действие:**" in line_s:
                current["next_action"] = line_s.split("**Следующее действие:**")[1].strip()
            elif "**Почему интересен:**" in line_s:
                current["why"] = line_s.split("**Почему интересен:**")[1].strip()
            elif "**Комментарий:**" in line_s:
                current["comment"] = line_s.split("**Комментарий:**")[1].strip()
            elif "**Пингануть:**" in line_s:
                date_str = line_s.split("**Пингануть:**")[1].strip()
                try:
                    current["ping_date"] = datetime.strptime(date_str, "%Y-%m-%d").date()
                except ValueError:
                    pass
    if current:
        contacts.append(current)
    return contacts


def format_crm_pretty(contacts: list[dict]) -> str:
    """Format CRM contacts for Telegram, highlighting overdue pings."""
    today = datetime.now().date()
    overdue = []
    upcoming = []
    no_date = []

    for c in contacts:
        if c["ping_date"]:
            if c["ping_date"] <= today:
                overdue.append(c)
            else:
                upcoming.append(c)
        else:
            no_date.append(c)

    overdue.sort(key=lambda x: x["ping_date"])
    upcoming.sort(key=lambda x: x["ping_date"])

    parts = []

    if overdue:
        parts.append("🔴 *Просроченные пинги:*")
        for c in overdue:
            days = (today - c["ping_date"]).days
            parts.append(
                f"  ⚠️ *{c['name']}* ({c['company']})\n"
                f"     Просрочено на {days} дн.\n"
                f"     → {c['next_action']}"
            )
        parts.append("")

    if upcoming:
        parts.append("🟡 *Предстоящие пинги:*")
        for c in upcoming:
            d = c["ping_date"]
            label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
            parts.append(
                f"  📅 *{c['name']}* ({c['company']}) — {label}\n"
                f"     → {c['next_action']}"
            )
        parts.append("")

    if no_date:
        parts.append("⚪ *Без даты пинга:*")
        for c in no_date:
            conf = f" [{c['confidence']}]" if c['confidence'] else ""
            parts.append(
                f"  • *{c['name']}* ({c['company']}){conf}\n"
                f"     → {c['next_action']}"
            )

    return "\n".join(parts) if parts else "CRM пуст"


def gpt_parse_crm_command(text: str) -> dict:
    """Use GPT to parse natural language CRM command."""
    prompt = f"""Пользователь хочет добавить или обновить контакт в CRM. Извлеки данные из текста.

Текст: {text}

Ответь в формате JSON (только JSON, без markdown):
{{
  "action": "add" или "update",
  "name": "Имя",
  "company": "Компания или пусто",
  "contact": "@телеграм или пусто",
  "warmth": "новый/холодный/средняя/Хороший/Хорошая или пусто",
  "confidence": "число от 0 до 100 или пусто",
  "next_action": "следующее действие или пусто",
  "why": "почему интересен или пусто",
  "comment": "комментарий или пусто",
  "ping_date": "YYYY-MM-DD или пусто"
}}

Если это обновление существующего контакта, укажи action=update и только изменённые поля (остальные пусто).
Если упоминается дата пинга (завтра, в понедельник, через 3 дня) — вычисли абсолютную дату от сегодня {datetime.now().strftime('%Y-%m-%d')} ({DAY_NAMES_SHORT[datetime.now().weekday()]})."""

    answer = claude_complete(prompt, max_tokens=300, temperature=0)
    answer = re.sub(r"^```(?:json)?\s*", "", answer)
    answer = re.sub(r"\s*```$", "", answer)
    return json.loads(answer)


def get_crm_file(project_id: str = None) -> str:
    """Get CRM file path for a project."""
    if project_id and project_id in PROJECTS:
        return PROJECTS[project_id].get("crm", CRM_FILE)
    return CRM_FILE


def find_contact_crm_file(name_query: str) -> str | None:
    """Find which CRM file contains a contact by name."""
    query_lower = name_query.lower()
    seen = set()
    for proj in PROJECTS.values():
        crm_file = proj.get("crm", CRM_FILE)
        if crm_file in seen:
            continue
        seen.add(crm_file)
        result = github_get_file(crm_file)
        if result:
            for line in result[0].split("\n"):
                if line.startswith("### "):
                    contact_name = line[4:].split("—")[0].strip().lower()
                    if query_lower in contact_name or contact_name in query_lower:
                        return crm_file
    return None


def add_crm_contact(data: dict, crm_file: str = None) -> str:
    """Add a new contact to CRM file."""
    target = crm_file or get_crm_file(data.get("project"))
    result = github_get_file(target)
    if not result:
        # Create new CRM file if it doesn't exist
        project_name = data.get("project", "")
        header = f"# CRM — {project_name}\n\n"
        content = header + "---\n*Last updated: " + datetime.now().strftime("%Y-%m-%d") + "*"
        github_put_file(target, content, f"crm: create {target}")
        result = github_get_file(target)
        if not result:
            return "CRM file error"
    content, sha = result

    name = data.get("name", "Без имени")
    company = data.get("company", "")
    header = f"### {name}" + (f" — {company}" if company else "")

    lines = [header]
    lines.append(f"- **Контакт:** {data.get('contact', '')}")
    warmth = data.get("warmth", "новый")
    conf = data.get("confidence", "0")
    lines.append(f"- **Теплота:** {warmth} | **Уверенность:** {conf}%")
    lines.append(f"- **Следующее действие:** {data.get('next_action', '')}")
    lines.append(f"- **Почему интересен:** {data.get('why', '')}")
    if data.get("comment"):
        lines.append(f"- **Комментарий:** {data['comment']}")
    if data.get("ping_date"):
        lines.append(f"- **Пингануть:** {data['ping_date']}")

    entry = "\n".join(lines)

    # Insert before the --- at the end
    content = content.replace("\n---\n*Last updated:", f"\n{entry}\n\n---\n*Last updated:")

    github_put_file(target, content, f"crm: add {name}", sha)
    return name


def update_crm_contact(name_query: str, data: dict, crm_file: str = None) -> str | None:
    """Update an existing contact in CRM by fuzzy name match."""
    # First try specified file, then search across all CRM files
    target = crm_file or find_contact_crm_file(name_query)
    if not target:
        target = get_crm_file(data.get("project"))
    result = github_get_file(target)
    if not result:
        return None
    content, sha = result
    contacts = parse_crm(content)

    # Fuzzy match
    query_lower = name_query.lower()
    matched = None
    for c in contacts:
        if query_lower in c["name"].lower() or c["name"].lower() in query_lower:
            matched = c
            break

    if not matched:
        return None

    # Rebuild the contact block
    old_block = "\n".join(matched["raw_lines"])
    new_lines = list(matched["raw_lines"])

    field_map = {
        "next_action": "**Следующее действие:**",
        "contact": "**Контакт:**",
        "comment": "**Комментарий:**",
        "why": "**Почему интересен:**",
    }

    for field, marker in field_map.items():
        val = data.get(field)
        if val:
            replaced = False
            for i, line in enumerate(new_lines):
                if marker in line:
                    new_lines[i] = f"- {marker} {val}"
                    replaced = True
                    break
            if not replaced:
                new_lines.append(f"- {marker} {val}")

    if data.get("warmth") or data.get("confidence"):
        for i, line in enumerate(new_lines):
            if "**Теплота:**" in line:
                w = data.get("warmth") or matched["warmth"]
                c = data.get("confidence", "").replace("%", "") or matched["confidence"].replace("%", "")
                new_lines[i] = f"- **Теплота:** {w} | **Уверенность:** {c}%"
                break

    if data.get("ping_date"):
        replaced = False
        for i, line in enumerate(new_lines):
            if "**Пингануть:**" in line:
                new_lines[i] = f"- **Пингануть:** {data['ping_date']}"
                replaced = True
                break
        if not replaced:
            new_lines.append(f"- **Пингануть:** {data['ping_date']}")

    new_block = "\n".join(new_lines)
    content = content.replace(old_block, new_block)
    github_put_file(target, content, f"crm: update {matched['name']}", sha)
    return matched["name"]


def gpt_parse_task_move(text: str) -> dict:
    """Use GPT to parse a task move command."""
    prompt = f"""Пользователь хочет перенести задачу на другую дату. Извлеки:
1. Ключевые слова задачи (по которым её найти)
2. Новая дата

Текст: {text}
Сегодня: {datetime.now().strftime('%Y-%m-%d')} ({DAY_NAMES_SHORT[datetime.now().weekday()]})

Ответь JSON (только JSON, без markdown):
{{
  "task_query": "ключевые слова для поиска задачи",
  "new_date": "YYYY-MM-DD"
}}"""

    answer = claude_complete(prompt, max_tokens=100, temperature=0)
    answer = re.sub(r"^```(?:json)?\s*", "", answer)
    answer = re.sub(r"\s*```$", "", answer)
    return json.loads(answer)


def move_task(task_query: str, new_date_str: str) -> str | None:
    """Find a task by fuzzy match and update its date."""
    from datetime import date as date_type
    try:
        new_date = datetime.strptime(new_date_str, "%Y-%m-%d").date()
    except ValueError:
        return None

    # Format label like "пн 09.03"
    new_label = f"{DAY_NAMES_SHORT[new_date.weekday()]} {new_date.strftime('%d.%m')}"

    # Search in backlog, weekly, daily
    for file_id in ["backlog.md", "weekly.md", "daily.md"]:
        result = github_get_file(file_id)
        if not result:
            continue
        content, sha = result
        lines = content.split("\n")
        query_lower = task_query.lower()

        for i, line in enumerate(lines):
            s = line.strip()
            if "[ ]" in s and query_lower in s.lower():
                # Remove old date suffix
                new_line = re.sub(r"\s*\(→[^)]+\)", "", line)
                # Add new date
                new_line = new_line.rstrip() + f" (→ {new_label})"
                lines[i] = new_line
                github_put_file(file_id, "\n".join(lines), f"move: {task_query[:30]} → {new_label}", sha)
                task_text = re.sub(r"^(?:-|\d+\.)\s*\[ \]\s*", "", s)
                task_text = re.sub(r"\s*\(→[^)]+\)", "", task_text)
                return f"{task_text} → {new_label}"

    return None


def detect_priority(text: str) -> tuple[str, str]:
    lower = text.lower()
    if re.search(URGENT_KEYWORDS, lower):
        clean = re.sub(URGENT_KEYWORDS, "", lower, count=1).strip(" ,:-—")
        return "urgent", clean or text
    if re.search(LOW_KEYWORDS, lower):
        clean = re.sub(LOW_KEYWORDS, "", lower, count=1).strip(" ,:-—")
        return "low", clean or text
    return "normal", text


def detect_project(text: str) -> str | None:
    lower = text.lower()
    for project_id, keywords in PROJECT_KEYWORDS.items():
        for kw in keywords:
            if kw in lower:
                return project_id
    return None


def detect_project_gpt(text: str, filename: str = "") -> str:
    prompt = f"""Определи к какому проекту относится этот документ/текст. Варианты:
- custody (криптокастодиан для банков, инвесторы, фандрайзинг, MPC)
- sber (стейблкоин Сбера, Гаймаков)
- reksoft (консалтинг по крипте через Reksoft, Скорочкин, Артём)
- blind-bets (проект Blind Bets, ставки, Денис, Autoforge)

Файл: {filename}
Текст: {text[:1000]}

Ответь ОДНИМ словом — id проекта (custody/sber/reksoft/blind-bets). Если не можешь определить — ответь unknown."""

    answer = claude_complete(prompt, max_tokens=10, temperature=0).lower()
    return answer if answer in PROJECTS else "unknown"


def summarize_document(text: str, filename: str = "") -> str:
    prompt = f"""Сделай краткое резюме документа на русском (3-5 пунктов). Выдели ключевые факты, цифры, действия.

Файл: {filename}
Текст:
{text[:4000]}"""

    return claude_complete(prompt, max_tokens=500, temperature=0.3)


INSIGHT_TAGS = [
    "финансы", "инвест", "продукт", "стратегия", "маркетинг",
    "технологии", "команда", "партнёры", "регуляция", "конкуренты",
    "продажи", "клиенты", "операции", "идея",
]


def detect_insight_tags(text: str) -> list[str]:
    """Use GPT to assign tags to an insight."""
    tags_list = ", ".join(INSIGHT_TAGS)
    prompt = f"""Прочитай инсайт и присвой ему 1-3 тега из списка: {tags_list}.

Инсайт: {text}

Ответь ТОЛЬКО тегами через запятую, без объяснений."""

    answer = claude_complete(prompt, max_tokens=30, temperature=0).lower()
    tags = [t.strip() for t in answer.split(",") if t.strip() in INSIGHT_TAGS]
    return tags if tags else ["идея"]


def save_link(project_id: str, url: str, description: str, tags: list[str]) -> None:
    """Save link to project's KB index.md under ## Полезные ссылки section."""
    project = PROJECTS[project_id]
    index_path = f"{project['path']}/index.md"
    result = github_get_file(index_path)
    if not result:
        return

    content, sha = result
    tags_str = " ".join(f"#{t}" for t in tags)
    title = description if description else url
    entry = f"- [{title}]({url}) `{tags_str}`"

    if "## Полезные ссылки" in content:
        content = content.replace("## Полезные ссылки", f"## Полезные ссылки\n{entry}", 1)
    elif "## Insights" in content:
        content = content.replace("## Insights", f"## Полезные ссылки\n{entry}\n\n## Insights")
    elif "## Documents" in content:
        content = content.replace("## Documents", f"## Полезные ссылки\n{entry}\n\n## Documents")

    github_put_file(index_path, content, f"link/{project_id}: {title[:40]}", sha)


def save_insight(project_id: str, text: str, tags: list[str]) -> None:
    """Save insight to project's KB index.md under ## Insights section."""
    project = PROJECTS[project_id]
    index_path = f"{project['path']}/index.md"
    result = github_get_file(index_path)
    if not result:
        return

    content, sha = result
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    tags_str = " ".join(f"#{t}" for t in tags)
    entry = f"- {text} — `{tags_str}` _{now}_"

    if "## Insights" in content:
        content = content.replace("## Insights", f"## Insights\n{entry}", 1)
    else:
        content = content.replace("## Documents", f"## Insights\n{entry}\n\n## Documents")

    github_put_file(index_path, content, f"insight/{project_id}: {text[:40]}", sha)


def split_into_tasks(text: str) -> list[str]:
    """Use GPT to split free-form text into actionable tasks."""
    prompt = f"""Пользователь написал сообщение в свободной форме. Разбей его на конкретные actionable задачи.

Правила:
- Каждая задача — одно конкретное действие (позвонить, написать, подготовить, etc.)
- Сохраняй имена, компании, детали из оригинала
- Если в тексте только одна задача — верни только её
- Если слова-маркеры приоритета (срочно, не важно, потом) — сохрани их в задаче
- Формат ответа: каждая задача на отдельной строке, без нумерации и маркеров

Сообщение: {text}"""

    response = claude_complete(prompt, max_tokens=300, temperature=0.2)
    tasks = [line.strip().lstrip("•-–123456789.) ") for line in response.split("\n") if line.strip()]
    return tasks if tasks else [text]


# --- Meeting parsing ---

def gpt_parse_meeting(notes: str, project_hint: str = "") -> dict:
    """Use GPT to parse meeting notes into structured entities."""
    projects_list = ", ".join(f"{k} ({v['name']})" for k, v in PROJECTS.items())
    prompt = f"""Ты разбираешь заметки со встречи на структурированные сущности.

Проекты: {projects_list}
{f"Подсказка по проекту: {project_hint}" if project_hint else ""}

Заметки:
{notes}

Верни JSON (без markdown-блоков):
{{
  "project": "project_id из списка выше",
  "title": "краткое название встречи (3-5 слов)",
  "decisions": ["список принятых решений/фактов"],
  "tasks": [
    {{"text": "конкретная задача", "assignee": "кто делает (если указано, иначе пусто)", "priority": "urgent/normal/low"}}
  ],
  "crm_updates": [
    {{"name": "Имя", "company": "Компания", "next_action": "следующее действие", "comment": "контекст"}}
  ]
}}

Правила:
- Каждая задача — конкретное действие (позвонить, написать, подготовить, подумать)
- Решения — это факты/договорённости, НЕ задачи
- Если упоминается человек с действием — это задача
- Если упоминается человек + компания с контекстом для запоминания — crm_updates
- assignee: если явно указано кто делает (напр. "Денис созвониться") — имя, иначе ""
- Приоритет: urgent если на этой неделе, normal — ближайшие 2 недели, low — потом
- crm_updates только если есть реально новый контакт или нужно обновить существующий
- Не дублируй одно и то же в decisions и tasks"""

    answer = claude_complete(prompt, max_tokens=1000, temperature=0.2)
    answer = re.sub(r"^```(?:json)?\s*", "", answer)
    answer = re.sub(r"\s*```$", "", answer)
    return json.loads(answer)


def smart_analyze(text: str) -> dict:
    """Use Claude to analyze free-form text and determine all needed actions."""
    projects_list = ", ".join(f"{k} ({v['name']})" for k, v in PROJECTS.items())
    today = datetime.now()
    date_str = today.strftime("%Y-%m-%d")
    weekday = DAY_NAMES_SHORT[today.weekday()]

    # Get CRM contact names for context (all CRM files)
    crm_names = []
    try:
        seen_crm = set()
        for proj in PROJECTS.values():
            crm_file = proj.get("crm", CRM_FILE)
            if crm_file in seen_crm:
                continue
            seen_crm.add(crm_file)
            result = github_get_file(crm_file)
            if result:
                for line in result[0].split("\n"):
                    if line.startswith("### "):
                        name = line.replace("### ", "").split(" — ")[0].strip()
                        crm_names.append(name)
    except Exception:
        pass
    crm_ctx = f"\nCRM контакты: {', '.join(crm_names)}" if crm_names else ""

    # Get open tasks for context (dedup + accurate closing)
    open_tasks = []
    try:
        for fid in ["daily.md", "backlog.md"]:
            result = github_get_file(fid)
            if result:
                for line in result[0].split("\n"):
                    s = line.strip()
                    if s.startswith("- [ ]") or re.match(r"^\d+\.\s*\[ \]", s):
                        t = re.sub(r"^(?:-|\d+\.)\s*\[ \]\s*", "", s)
                        if t and t != "[Task]":
                            open_tasks.append(t)
    except Exception:
        pass
    tasks_ctx = f"\nОткрытые задачи:\n" + "\n".join(f"- {t}" for t in open_tasks[:30]) if open_tasks else ""

    prompt = f"""Ты — ассистент CEO крипто-стартапа. Проанализируй сообщение и определи ВСЕ действия.

Сегодня: {date_str} ({weekday})
Проекты: {projects_list}{crm_ctx}{tasks_ctx}

Сообщение:
{text}

Верни JSON (без markdown):
{{
  "reply": "текст ответа если пользователь задаёт ВОПРОС (не задачу). Иначе null",
  "tasks": [
    {{"text": "конкретная задача", "priority": "urgent/normal/low", "date": "YYYY-MM-DD или пусто"}}
  ],
  "crm_updates": [
    {{"name": "Имя", "company": "Компания или пусто", "project": "project_id", "action": "add/update", "next_action": "следующее действие или пусто", "comment": "комментарий или пусто", "ping_date": "YYYY-MM-DD или пусто", "warmth": "пусто или новый/холодный/средняя/Хороший"}}
  ],
  "insights": [
    {{"text": "инсайт/наблюдение/факт", "project": "project_id"}}
  ],
  "done": [
    {{"query": "ключевые слова задачи для закрытия"}}
  ],
  "moves": [
    {{"query": "ключевые слова задачи", "new_date": "YYYY-MM-DD"}}
  ],
  "meeting": null
}}

Правила:
- ВОПРОСЫ: если пользователь спрашивает вопрос ("в какой проект?", "что у меня на завтра?", "сколько задач?") — верни ТОЛЬКО "reply" с ответом. НЕ создавай задачи из вопросов!
- Если простое действие (позвонить, написать, подготовить) — tasks
- Если упоминается человек + новая инфо для запоминания — crm_updates
- Если наблюдение/вывод о рынке/продукте/стратегии — insights
- Если "сделал/готово/закрыл/отправил" что-то — done (найдём задачу по ключевым словам)
- Если "перенеси/сдвинь на..." — moves
- Если заметки со встречи (несколько решений, разные люди, план) — meeting с полными заметками
- ОДНО сообщение может содержать НЕСКОЛЬКО типов действий одновременно
- ПРИОРИТЕТ: задача на эту неделю (до 7 дней) = "urgent". Задача на 2+ недели = "normal". Без дедлайна = "low"
- date: "завтра" = +1 день, "в понедельник" = ближайший пн, "на следующей неделе" = пн след. недели
- Если action=update для CRM и контакт уже в списке — обнови, иначе action=add
- CRM project: определи к какому проекту относится контакт (custody, sber, blind-bets, reksoft). Гаймаков/Сбер = "sber"
- Пустые массивы НЕ включай в ответ
- Если не уверен insight это или задача — делай задачу
- Для meeting: передай ВСЕ заметки в notes целиком
- ДУБЛИКАТЫ: если в открытых задачах уже есть похожая — НЕ создавай новую, а если нужно обновить — используй done для старой + новый task
- done.query: используй ключевые слова из ОТКРЫТЫХ ЗАДАЧ выше (имя, компанию) чтобы точно найти"""

    answer = claude_complete(prompt, max_tokens=1000, temperature=0.1)
    answer = re.sub(r"^```(?:json)?\s*", "", answer)
    answer = re.sub(r"\s*```$", "", answer)
    return json.loads(answer)


def save_meeting_log(project_id: str, title: str, raw_notes: str, parsed: dict) -> str:
    """Save meeting log to meetings/ folder and add reference to project index."""
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H-%M")
    safe_title = re.sub(r"[^\w\-]", "_", title)[:40]

    decisions_md = "\n".join(f"- {d}" for d in parsed.get("decisions", [])) or "---"
    tasks_md = "\n".join(
        f"- [ ] {t['text']}" + (f" (@{t['assignee']})" if t.get("assignee") else "")
        for t in parsed.get("tasks", [])
    ) or "---"
    crm_md = "\n".join(
        f"- {c['name']}" + (f" ({c.get('company', '')})" if c.get("company") else "") +
        f" -> {c.get('next_action', '')}"
        for c in parsed.get("crm_updates", [])
    ) or "---"

    content = (
        f"# {title}\n"
        f"*Date: {date_str}*\n"
        f"*Project: {PROJECTS[project_id]['name']}*\n\n"
        f"## Raw Notes\n{raw_notes}\n\n"
        f"## Decisions\n{decisions_md}\n\n"
        f"## Tasks\n{tasks_md}\n\n"
        f"## CRM Updates\n{crm_md}\n"
    )

    meeting_path = f"meetings/{date_str}_{time_str}_{safe_title}.md"
    github_put_file(meeting_path, content, f"meeting/{project_id}: {title[:40]}")

    # Add meeting reference to project index
    project = PROJECTS[project_id]
    index_path = f"{project['path']}/index.md"
    result = github_get_file(index_path)
    if result:
        idx_content, sha = result
        meeting_entry = f"- [{title}](../../{meeting_path}) --- _{date_str}_"
        if "## Meetings" in idx_content:
            idx_content = idx_content.replace("## Meetings", f"## Meetings\n{meeting_entry}", 1)
        elif "## Documents" in idx_content:
            idx_content = idx_content.replace("## Documents", f"## Meetings\n{meeting_entry}\n\n## Documents")
        else:
            idx_content += f"\n\n## Meetings\n{meeting_entry}"
        github_put_file(index_path, idx_content, f"meeting-ref/{project_id}", sha)

    return meeting_path


# --- GitHub helpers ---

def github_get_file(path: str) -> tuple[str, str] | None:
    r = httpx.get(f"{GITHUB_BASE}/{path}", headers=GH_HEADERS)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    data = r.json()
    content = base64.b64decode(data["content"]).decode("utf-8")
    return content, data["sha"]


def github_put_file(path: str, content: str, message: str, sha: str = None) -> None:
    encoded = base64.b64encode(content.encode("utf-8")).decode("utf-8")
    payload = {"message": message, "content": encoded}
    if sha:
        payload["sha"] = sha
    r = httpx.put(f"{GITHUB_BASE}/{path}", headers=GH_HEADERS, json=payload)
    if r.status_code == 409 and sha:
        # SHA conflict — re-read and retry (someone else wrote in between)
        print(f"409 conflict on {path}, retrying with fresh SHA...", flush=True)
        fresh = github_get_file(path)
        if fresh:
            payload["sha"] = fresh[1]
            r = httpx.put(f"{GITHUB_BASE}/{path}", headers=GH_HEADERS, json=payload)
    r.raise_for_status()


def add_task_to_github(task_text: str) -> str:
    date_label, clean_text = parse_date(task_text)
    priority, clean_text = detect_priority(clean_text)
    section = SECTION_MAP[priority]

    result = github_get_file(BACKLOG_FILE)
    if not result:
        return "error"
    content, sha = result

    date_suffix = f" (→ {date_label})" if date_label else ""
    task_line = f"- [ ] {clean_text}{date_suffix}"
    lines = content.split("\n")
    out = []
    inserted = False
    in_section = False

    for line in lines:
        if line.strip() == section:
            in_section = True
            out.append(line)
            continue
        if in_section:
            if line.startswith("- [ ]"):
                out.append(line)
                continue
            else:
                out.append(task_line)
                inserted = True
                in_section = False
                out.append(line)
                continue
        out.append(line)

    if not inserted:
        out.append(task_line)

    github_put_file(BACKLOG_FILE, "\n".join(out), f"backlog: {task_text[:50]}", sha)
    return PRIORITY_LABELS[priority]


def save_to_kb(project_id: str, filename: str, text: str, summary: str, file_id: str = "") -> None:
    project = PROJECTS[project_id]
    now = datetime.now().strftime("%Y-%m-%d_%H-%M")
    safe_name = re.sub(r"[^\w\-.]", "_", filename.rsplit(".", 1)[0])
    doc_path = f"{project['path']}/{now}_{safe_name}.md"

    file_id_line = f"*File ID: {file_id}*\n" if file_id else ""
    doc_content = f"""# {filename}
*Added: {datetime.now().strftime("%Y-%m-%d %H:%M")}*
*Project: {project['name']}*
{file_id_line}
## Summary
{summary}

## Full Text
{text[:10000]}
"""
    github_put_file(doc_path, doc_content, f"kb/{project_id}: {filename[:40]}")

    # Update index
    result = github_get_file(f"{project['path']}/index.md")
    if result:
        index_content, index_sha = result
        entry = f"- [{filename}]({now}_{safe_name}.md) — {datetime.now().strftime('%Y-%m-%d')}"
        index_content = index_content.replace(
            "## Documents",
            f"## Documents\n{entry}",
        )
        github_put_file(
            f"{project['path']}/index.md",
            index_content,
            f"kb index: {filename[:40]}",
            index_sha,
        )


def collect_open_tasks() -> list[dict]:
    """Collect all open tasks from daily, weekly, backlog."""
    tasks = []
    for file_id, label in [("daily.md", "today"), ("weekly.md", "week"), ("backlog.md", "backlog")]:
        result = github_get_file(file_id)
        if not result:
            continue
        content, _ = result
        for i, line in enumerate(content.split("\n")):
            s = line.strip()
            if s.startswith("- [ ]") or re.match(r"^\d+\.\s*\[ \]", s):
                task_text = re.sub(r"^(?:-|\d+\.)\s*\[ \]\s*", "", s)
                if task_text and task_text != "[Task]":
                    tasks.append({"file": file_id, "line_idx": i, "text": task_text, "source": label})
    return tasks


def toggle_task_in_file(file_id: str, line_idx: int, close: bool = True) -> bool:
    result = github_get_file(file_id)
    if not result:
        return False
    content, sha = result
    lines = content.split("\n")
    if line_idx >= len(lines):
        return False

    old = "[ ]" if close else "[x]"
    new = "[x]" if close else "[ ]"
    if old in lines[line_idx]:
        lines[line_idx] = lines[line_idx].replace(old, new, 1)
        github_put_file(file_id, "\n".join(lines), f"{'close' if close else 'reopen'}: {lines[line_idx][:50]}", sha)
        return True
    return False


# --- Telegram handlers ---

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    help_text = (
        "Привет! Я твой личный ассистент.\n\n"

        "*ПРОСТО ПИШИ — Я ПОЙМУ*\n"
        "Пиши текстом или голосом в свободной форме.\n"
        "Я сам определю что делать:\n"
        "- задача → backlog\n"
        "- человек + инфо → CRM\n"
        "- наблюдение/факт → insight в KB\n"
        "- заметки встречи → лог + задачи + CRM\n"
        "- «сделал/готово» → закрою задачу\n"
        "- «перенеси на...» → сдвину дату\n\n"

        "*Примеры:*\n"
        "`завтра написать Грише из ВТБ`\n"
        "`Созвонился с Колей из Азуро, готов обсуждать ЦА`\n"
        "`отправил описание Шмакову, пингануть в понедельник`\n"
        "`встреча blind bets: 1 раунд 8 часов, ZK депозиты`\n\n"

        "*КОМАНДЫ*\n"
        "/today /week /backlog /crm /meetings\n"
        "/done /insights /project /search\n\n"

        "*ШОРТКАТЫ (быстрее, без AI)*\n"
        "`готово шмаков` — закрыть задачу\n"
        "`перенеси Гришу на среду` — сдвинуть\n"
        "`ссылка: https://... описание` — в KB\n\n"

        "📎 PDF/XLSX/DOCX — отправь файл, сохраню в KB\n"
        "🎤 Голос — всё работает голосом\n\n"
        "⏰ *УТРЕННЕЕ НАПОМИНАНИЕ*\n"
        "Каждый день в 9:00 МСК — задачи на сегодня\n"
        "+ просроченные пинги CRM"
    )
    await update.message.reply_text(help_text, parse_mode="Markdown")


def escape_md(text: str) -> str:
    """Escape Markdown special characters for Telegram."""
    for ch in ("_", "*", "`", "[", "]", "(", ")", "~", ">", "#", "+", "-", "=", "|", "{", "}", ".", "!"):
        text = text.replace(ch, f"\\{ch}")
    return text


def format_file_pretty(file_id: str, content: str) -> tuple[str, list[dict]]:
    """Format markdown nicely for Telegram, collect open tasks."""
    lines = content.split("\n")
    formatted = []
    tasks = []

    for i, line in enumerate(lines):
        s = line.strip()
        if not s or s.startswith("---") or s.startswith("*Last updated"):
            continue
        if s.startswith("# "):
            formatted.append(f"*{s[2:]}*\n")
        elif s.startswith("## "):
            formatted.append(f"📌 *{s[3:]}*")
        elif s.startswith("### "):
            formatted.append(f"  🔹 *{s[4:]}*")
        elif s.startswith("> "):
            formatted.append(f"_{s[2:]}_")
        elif s.startswith("- [x]") or re.match(r"^\d+\.\s*\[x\]", s):
            text_part = re.sub(r"^(?:-|\d+\.)\s*\[x\]\s*", "", s)
            formatted.append(f"  ✅ ~{text_part}~")
        elif s.startswith("- [ ]") or re.match(r"^\d+\.\s*\[ \]", s):
            task_text = re.sub(r"^(?:-|\d+\.)\s*\[ \]\s*", "", s)
            if task_text and task_text != "[Task]":
                num = len(tasks) + 1
                formatted.append(f"  ⬜ {num}. {task_text}")
                tasks.append({"file": file_id, "line_idx": i, "text": task_text})
        elif s.startswith("- **"):
            formatted.append(f"  {s[2:]}")
        elif s.startswith("|"):
            formatted.append(f"`{s}`")
        else:
            formatted.append(s)

    return "\n".join(formatted), tasks


def build_task_buttons(tasks: list[dict], prefix: str = "close") -> InlineKeyboardMarkup | None:
    if not tasks:
        return None
    keyboard = []
    row = []
    for i, t in enumerate(tasks[:20]):
        row.append(InlineKeyboardButton(f"✅ {i+1}", callback_data=f"{prefix}:{t['file']}:{t['line_idx']}"))
        if len(row) == 5:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    return InlineKeyboardMarkup(keyboard)


async def send_formatted(update, file_id: str, context) -> None:
    result = github_get_file(file_id)
    if not result:
        await update.message.reply_text(f"{file_id} не найден")
        return
    text, tasks = format_file_pretty(file_id, result[0])
    markup = build_task_buttons(tasks)
    await update.message.reply_text(text[:4000], parse_mode="Markdown", reply_markup=markup)


async def cmd_today(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    await send_formatted(update, "daily.md", context)


async def cmd_week(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    await send_formatted(update, "weekly.md", context)


async def cmd_backlog(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    await send_formatted(update, "backlog.md", context)


async def cmd_crm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    # Show all CRM files
    all_parts = []
    seen = set()
    for pid, proj in PROJECTS.items():
        crm_file = proj.get("crm", CRM_FILE)
        if crm_file in seen:
            continue
        seen.add(crm_file)
        result = github_get_file(crm_file)
        if result:
            contacts = parse_crm(result[0])
            if contacts:
                text = format_crm_pretty(contacts)
                all_parts.append(text)
    if all_parts:
        combined = "\n\n".join(all_parts)
        await update.message.reply_text(combined[:4000], parse_mode="Markdown")
    else:
        await update.message.reply_text("CRM пуст")


async def cmd_meetings(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    try:
        r = httpx.get(f"{GITHUB_BASE}/meetings", headers=GH_HEADERS, timeout=10)
        if r.status_code == 404:
            await update.message.reply_text("Пока нет записей встреч.")
            return
        r.raise_for_status()
        files = r.json()
        md_files = [f for f in files if f["name"].endswith(".md")]
        md_files.sort(key=lambda f: f["name"], reverse=True)

        if not md_files:
            await update.message.reply_text("Пока нет записей встреч.")
            return

        lines = ["*Последние встречи:*\n"]
        for f in md_files[:10]:
            name = f["name"].replace(".md", "").replace("_", " ")
            lines.append(f"- `{name}`")

        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"Ошибка: {e}")


async def cmd_morning(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    try:
        await update.message.reply_text("Генерирую брифинг...")
        ctx = collect_morning_context()
        briefing = generate_morning_briefing(ctx)
        await update.message.reply_text(briefing[:4000], parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"Ошибка: {e}")


async def cmd_evening(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    text = " ".join(context.args) if context.args else ""
    if not text:
        await update.message.reply_text(
            "Расскажи как прошёл день:\n"
            "`/evening отправил Шмакову, созвонился с Колей, не успел Никиту, покурил кальян`"
            , parse_mode="Markdown")
        return
    await process_evening(update, text)


async def process_evening(update: Update, text: str) -> None:
    """Process evening review: close tasks, update CRM, journal, habits."""
    await update.message.reply_text("Анализирую день...")
    try:
        data = evening_analyze(text)
        parts = []

        # 1. Close done tasks
        for d in data.get("done", []):
            try:
                all_tasks = collect_open_tasks()
                q = d["query"].lower()
                matched = [t for t in all_tasks if q in t["text"].lower()]
                if matched:
                    ok = toggle_task_in_file(matched[0]["file"], matched[0]["line_idx"], close=True)
                    if ok:
                        parts.append(f"Done: ~{matched[0]['text']}~")
            except Exception:
                pass

        # 2. CRM updates (auto-detect CRM file)
        for c in data.get("crm_updates", []):
            try:
                name = update_crm_contact(c.get("name", ""), c)
                if name:
                    parts.append(f"CRM: *{name}* — {c.get('comment', '')}")
            except Exception:
                pass

        # 3. New tasks
        new_tasks = data.get("new_tasks", [])
        if new_tasks:
            backlog_result = github_get_file(BACKLOG_FILE)
            if backlog_result:
                content, sha = backlog_result
                for t in new_tasks:
                    task_line = f"- [ ] {t['text']}"
                    section = SECTION_MAP.get(t.get("priority", "normal"), SECTION_MAP["normal"])
                    if section not in content:
                        section = SECTION_MAP["normal"]
                    lines = content.split("\n")
                    out = []
                    inserted = False
                    in_section = False
                    for line in lines:
                        if line.strip() == section:
                            in_section = True
                            out.append(line)
                            continue
                        if in_section:
                            if line.startswith("- [ ]") or line.startswith("- [x]"):
                                out.append(line)
                                continue
                            else:
                                out.append(task_line)
                                inserted = True
                                in_section = False
                        out.append(line)
                    if not inserted:
                        out.append(task_line)
                    content = "\n".join(out)
                    parts.append(f"New: {t['text']}")
                github_put_file(BACKLOG_FILE, content, f"evening-tasks", sha)

        # 4. Habits
        habits_data = data.get("habits")
        if habits_data:
            update_habit_tracker(habits_data)
            done_h = habits_data.get("done", [])
            skip_h = habits_data.get("skipped", [])
            if done_h:
                parts.append(f"Habits done: {', '.join(done_h)}")
            if skip_h:
                parts.append(f"Habits skipped: {', '.join(skip_h)}")

        # 5. Journal
        today_str = datetime.now().strftime("%Y-%m-%d")
        save_journal_entry(today_str, data)
        parts.append(f"Journal: `journal/{today_str}.md`")

        # 6. Summary
        summary = data.get("summary", {})
        if summary.get("mood"):
            parts.insert(0, f"*Day: {summary['mood']}*")

        await update.message.reply_text("\n".join(parts), parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"Ошибка: {e}")


async def cmd_done(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    tasks = collect_open_tasks()
    if not tasks:
        await update.message.reply_text("Нет открытых задач!")
        return

    # Store tasks in context for callback
    context.user_data["tasks"] = tasks

    # Build inline keyboard (max 30 tasks to avoid Telegram limits)
    keyboard = []
    for i, t in enumerate(tasks[:30]):
        icon = {"today": "📅", "week": "📋", "backlog": "📝"}.get(t["source"], "")
        label = f"{icon} {t['text'][:45]}"
        keyboard.append([InlineKeyboardButton(label, callback_data=f"done:{i}")])

    await update.message.reply_text(
        "Выбери задачу для закрытия:",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if query.from_user.id != ALLOWED_USER_ID:
        return
    await query.answer()

    data = query.data

    if data.startswith("refresh:"):
        file_id = data.split(":", 1)[1]
        result = github_get_file(file_id)
        if result:
            text, tasks = format_file_pretty(file_id, result[0])
            markup = build_task_buttons(tasks)
            await query.edit_message_text(text[:4000], parse_mode="Markdown", reply_markup=markup)
        else:
            await query.edit_message_text(f"{file_id} не найден")

    elif data.startswith("close:"):
        parts = data.split(":", 2)
        file_id = parts[1]
        line_idx = int(parts[2])
        ok = toggle_task_in_file(file_id, line_idx, close=True)
        if ok:
            # Read closed task text for confirmation
            result = github_get_file(file_id)
            closed_text = ""
            if result:
                lines = result[0].split("\n")
                if line_idx < len(lines):
                    closed_text = re.sub(r"^(?:-|\d+\.)\s*\[x\]\s*", "", lines[line_idx].strip())
            refresh_btn = InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Обновить список", callback_data=f"refresh:{file_id}")]
            ])
            await query.edit_message_text(
                f"✅ {closed_text}\n\nНажми «Обновить» чтобы увидеть остальные задачи.",
                reply_markup=refresh_btn,
            )
        else:
            await query.edit_message_text("Не удалось закрыть задачу.")

    elif data.startswith("done:"):
        idx = int(data.split(":")[1])
        tasks = context.user_data.get("tasks", [])
        if idx >= len(tasks):
            await query.edit_message_text("Задача не найдена.")
            return
        task = tasks[idx]
        ok = toggle_task_in_file(task["file"], task["line_idx"], close=True)
        if ok:
            await query.edit_message_text(f"✅ Закрыто:\n~{task['text']}~", parse_mode="Markdown")
        else:
            await query.edit_message_text("Не удалось закрыть задачу.")


async def cmd_insights(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    args = context.args
    project_ids = [args[0]] if args and args[0] in PROJECTS else list(PROJECTS.keys())
    parts = []
    for pid in project_ids:
        result = github_get_file(f"{PROJECTS[pid]['path']}/index.md")
        if not result:
            continue
        content = result[0]
        if "## Insights" not in content:
            continue
        # Extract insights section
        section = content.split("## Insights")[1].split("## ")[0].strip()
        if section:
            parts.append(f"💡 *{PROJECTS[pid]['name']}*\n{section}")
    if parts:
        await update.message.reply_text("\n\n".join(parts)[:4000], parse_mode="Markdown")
    else:
        await update.message.reply_text("Пока нет инсайтов. Напиши `инсайт: твоя мысль`", parse_mode="Markdown")


def extract_doc_info(project_path: str, doc_link: str) -> dict | None:
    """Fetch a KB document and extract its file_id and summary."""
    m = re.search(r"\(([^)]+\.md)\)", doc_link)
    if not m:
        return None
    doc_path = f"{project_path}/{m.group(1)}"
    result = github_get_file(doc_path)
    if not result:
        return None
    content = result[0]
    info = {}
    # Extract file_id
    fid_m = re.search(r"\*File ID:\s*([^\*]+)\*", content)
    if fid_m:
        info["file_id"] = fid_m.group(1).strip()
    # Extract summary
    if "## Summary" in content:
        info["summary"] = content.split("## Summary")[1].split("\n## ")[0].strip()
    # Extract original filename from title
    title_m = re.match(r"# (.+)", content)
    if title_m:
        info["filename"] = title_m.group(1).strip()
    return info if info else None


async def cmd_project(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    args = context.args
    if not args or args[0] not in PROJECTS:
        keys = ", ".join(f"`{k}`" for k in PROJECTS)
        await update.message.reply_text(f"Укажи проект: /project {keys}", parse_mode="Markdown")
        return
    pid = args[0]
    project = PROJECTS[pid]
    result = github_get_file(f"{project['path']}/index.md")
    if not result:
        await update.message.reply_text("index.md не найден")
        return
    content = result[0]

    # Extract "О проекте" section
    about = None
    if "## О проекте" in content:
        about = content.split("## О проекте")[1].split("\n## ")[0].strip()

    # Extract key doc links
    key_docs = []
    if "**Ключевые документы**" in content:
        docs_section = content.split("**Ключевые документы**")[1].split("\n## ")[0].strip()
        for line in docs_section.split("\n"):
            line = line.strip()
            if line.startswith("- ["):
                title_m = re.match(r"- \[([^\]]+)\]", line)
                title = title_m.group(1) if title_m else "Документ"
                info = extract_doc_info(project["path"], line)
                if info:
                    info["title"] = title
                    key_docs.append(info)

    await update.message.reply_text(f"📋 Загружаю данные по {project['name']}...")

    # Send project description
    parts = [f"📋 *{project['name']}*"]
    if about:
        parts.append(about.replace("**", "*"))
    else:
        parts.append("Описание проекта пока не заполнено.")
    await update.message.reply_text("\n\n".join(parts)[:4000], parse_mode="Markdown")

    # Send key documents
    bot = update.get_bot()
    chat_id = update.effective_chat.id
    for doc in key_docs:
        if doc.get("file_id"):
            # Send original file
            try:
                await bot.send_document(chat_id=chat_id, document=doc["file_id"])
                continue
            except Exception:
                pass
        # Fallback: send summary as text
        clean_title = doc.get("title", "Документ").replace("_", " ")
        summary = doc.get("summary", "Нет саммари")
        clean_summary = summary.replace("*", "").replace("`", "").replace("_", " ")
        await update.message.reply_text(
            f"📎 *{clean_title}*\n\n{clean_summary}"[:4000],
            parse_mode="Markdown",
        )


async def cmd_tracks(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    await send_formatted(update, "tracks.md", context)


async def transcribe_voice(voice_file) -> str:
    buf = bytearray()
    await voice_file.download_as_bytearray(buf)
    r = httpx.post(
        "https://api.openai.com/v1/audio/transcriptions",
        headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
        data={"model": "whisper-1", "language": "ru"},
        files={"file": ("voice.ogg", bytes(buf), "audio/ogg")},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()["text"]


async def process_task(update: Update, task_text: str, source: str = "") -> None:
    try:
        # Extract date BEFORE GPT splitting (GPT may rephrase date words)
        date_label_orig, _ = parse_date(task_text)

        tasks = split_into_tasks(task_text)
        prefix = "🎤 " if source == "voice" else ""
        results = []
        for task in tasks:
            # Try parsing date from the split task first
            date_label, clean = parse_date(task)
            if not date_label and date_label_orig:
                # GPT lost the date — use original
                date_label = date_label_orig
                clean = task

            if date_label:
                priority, clean = detect_priority(clean)
                section = SECTION_MAP[priority]
                date_suffix = f" (→ {date_label})"
                task_line = f"- [ ] {clean}{date_suffix}"

                result = github_get_file(BACKLOG_FILE)
                if result:
                    content, sha = result
                    lines = content.split("\n")
                    out = []
                    inserted = False
                    in_section = False
                    for line in lines:
                        if line.strip() == section:
                            in_section = True
                            out.append(line)
                            continue
                        if in_section:
                            if line.startswith("- [ ]"):
                                out.append(line)
                                continue
                            else:
                                out.append(task_line)
                                inserted = True
                                in_section = False
                                out.append(line)
                                continue
                        out.append(line)
                    if not inserted:
                        out.append(task_line)
                    github_put_file(BACKLOG_FILE, "\n".join(out), f"backlog: {clean[:50]}", sha)
                    label = PRIORITY_LABELS[priority]
                else:
                    label = "error"
                results.append(f"[{label}] `{clean}` 📅 {date_label}")
            else:
                label = add_task_to_github(task)
                results.append(f"[{label}] `{task}`")
        msg = prefix + "Добавлено в backlog:\n" + "\n".join(results)
        await update.message.reply_text(msg, parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"Ошибка: {e}")


async def process_meeting(update: Update, notes: str, project_hint: str = "") -> None:
    """Process meeting notes: parse, save log, create tasks, update CRM."""
    await update.message.reply_text("Разбираю заметки встречи...")

    try:
        parsed = gpt_parse_meeting(notes, project_hint)
        project_id = parsed.get("project", "")
        if project_id not in PROJECTS:
            project_id = detect_project(notes) or detect_project_gpt(notes)
            if project_id not in PROJECTS:
                project_id = list(PROJECTS.keys())[0]

        title = parsed.get("title", "Meeting")

        # 1. Save meeting log
        meeting_path = save_meeting_log(project_id, title, notes, parsed)

        # 2. Add tasks to backlog (batch — single read/write to avoid SHA conflict)
        tasks = parsed.get("tasks", [])
        task_results = []
        if tasks:
            result = github_get_file(BACKLOG_FILE)
            if result:
                content, sha = result
                for t in tasks:
                    task_text = t["text"]
                    if t.get("assignee"):
                        task_text = f"{t['assignee']}: {task_text}"
                    priority_key = {"urgent": "urgent", "low": "low"}.get(t.get("priority", ""), "normal")
                    section = SECTION_MAP[priority_key]
                    # Fall back to Important if target section doesn't exist
                    if section not in content:
                        section = SECTION_MAP["normal"]
                    task_line = f"- [ ] {task_text}"
                    lines = content.split("\n")
                    out = []
                    inserted = False
                    in_section = False
                    for line in lines:
                        if line.strip() == section:
                            in_section = True
                            out.append(line)
                            continue
                        if in_section:
                            if line.startswith("- [ ]") or line.startswith("- [x]"):
                                out.append(line)
                                continue
                            else:
                                out.append(task_line)
                                inserted = True
                                in_section = False
                        out.append(line)
                    if not inserted:
                        out.append(task_line)
                    content = "\n".join(out)
                    task_results.append(f"  [{PRIORITY_LABELS[priority_key]}] {task_text}")
                github_put_file(BACKLOG_FILE, content, f"meeting-tasks: {title[:40]}", sha)

        # 3. Update CRM (use project-specific CRM file)
        crm_target = get_crm_file(project_id)
        crm_results = []
        for c in parsed.get("crm_updates", []):
            try:
                name = update_crm_contact(c.get("name", ""), c, crm_target)
                if name:
                    crm_results.append(f"  upd {name}")
                else:
                    name = add_crm_contact(c, crm_target)
                    crm_results.append(f"  + {name}")
            except Exception:
                pass

        # 4. Build summary
        decisions = parsed.get("decisions", [])
        parts = [f"*Meeting: {title}*", f"Project: {PROJECTS[project_id]['name']}"]

        if decisions:
            parts.append("\n*Decisions:*")
            for d in decisions:
                parts.append(f"  > {d}")

        if task_results:
            parts.append(f"\n*Tasks ({len(task_results)}):*")
            parts.extend(task_results)

        if crm_results:
            parts.append(f"\n*CRM:*")
            parts.extend(crm_results)

        parts.append(f"\nLog: `{meeting_path}`")

        await update.message.reply_text("\n".join(parts), parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"Ошибка обработки встречи: {e}")


async def smart_process(update: Update, text: str, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Intelligently process free-form text into multiple actions."""
    try:
        result = smart_analyze(text)
        parts = []

        # 0. Reply — answer a question without creating tasks
        reply = result.get("reply")
        if reply and not result.get("tasks") and not result.get("crm_updates") and not result.get("done"):
            await update.message.reply_text(reply)
            return

        # 1. Tasks — batch write to backlog
        tasks = result.get("tasks", [])
        if tasks:
            backlog_result = github_get_file(BACKLOG_FILE)
            if backlog_result:
                content, sha = backlog_result
                for t in tasks:
                    priority_key = {"urgent": "urgent", "low": "low"}.get(t.get("priority", ""), "normal")
                    section = SECTION_MAP[priority_key]
                    if section not in content:
                        section = SECTION_MAP["normal"]
                    date_str = t.get("date", "")
                    if date_str:
                        from datetime import date as dt_date
                        try:
                            d = dt_date.fromisoformat(date_str)
                            date_label = f"{DAY_NAMES_SHORT[d.weekday()]} {d.strftime('%d.%m')}"
                        except ValueError:
                            date_label = date_str
                        date_suffix = f" (→ {date_label})"
                    else:
                        date_suffix = ""
                    task_line = f"- [ ] {t['text']}{date_suffix}"
                    lines = content.split("\n")
                    out = []
                    inserted = False
                    in_section = False
                    for line in lines:
                        if line.strip() == section:
                            in_section = True
                            out.append(line)
                            continue
                        if in_section:
                            if line.startswith("- [ ]") or line.startswith("- [x]"):
                                out.append(line)
                                continue
                            else:
                                out.append(task_line)
                                inserted = True
                                in_section = False
                        out.append(line)
                    if not inserted:
                        out.append(task_line)
                    content = "\n".join(out)
                    label = PRIORITY_LABELS[priority_key]
                    parts.append(f"[{label}] {t['text']}{date_suffix}")
                github_put_file(BACKLOG_FILE, content, f"smart: {tasks[0]['text'][:40]}", sha)

        # 2. Done — close tasks
        for d in result.get("done", []):
            try:
                all_tasks = collect_open_tasks()
                q = d["query"].lower()
                matched = [t for t in all_tasks if q in t["text"].lower()]
                if matched:
                    ok = toggle_task_in_file(matched[0]["file"], matched[0]["line_idx"], close=True)
                    if ok:
                        parts.append(f"Done: ~{matched[0]['text']}~")
            except Exception:
                pass

        # 3. Moves
        for m in result.get("moves", []):
            try:
                moved = move_task(m["query"], m["new_date"])
                if moved:
                    parts.append(f"Moved: {moved}")
            except Exception:
                pass

        # 4. CRM updates (project-aware)
        for c in result.get("crm_updates", []):
            try:
                project_id = c.get("project", "")
                crm_target = get_crm_file(project_id) if project_id else None
                if c.get("action") == "update":
                    name = update_crm_contact(c.get("name", ""), c, crm_target)
                    if name:
                        detail = c.get("next_action") or c.get("comment") or ""
                        proj_label = PROJECTS[project_id]["name"] if project_id in PROJECTS else ""
                        parts.append(f"CRM upd: *{name}*" + (f" ({proj_label})" if proj_label else "") + (f" → {detail}" if detail else ""))
                    else:
                        name = add_crm_contact(c, crm_target)
                        proj_label = PROJECTS[project_id]["name"] if project_id in PROJECTS else ""
                        parts.append(f"CRM +: *{name}*" + (f" ({proj_label})" if proj_label else ""))
                else:
                    name = add_crm_contact(c, crm_target)
                    proj_label = PROJECTS[project_id]["name"] if project_id in PROJECTS else ""
                    parts.append(f"CRM +: *{name}*" + (f" ({proj_label})" if proj_label else ""))
            except Exception:
                pass

        # 5. Insights
        for ins in result.get("insights", []):
            try:
                pid = ins.get("project", "")
                if pid not in PROJECTS:
                    pid = detect_project(ins["text"]) or "custody"
                tags = detect_insight_tags(ins["text"])
                save_insight(pid, ins["text"], tags)
                tags_str = " ".join(f"#{t}" for t in tags)
                parts.append(f"Insight: {ins['text'][:80]} {tags_str}")
            except Exception:
                pass

        # 6. Meeting — delegate to process_meeting
        meeting = result.get("meeting")
        if meeting:
            notes = meeting.get("notes", text)
            hint = meeting.get("project_hint", "")
            await process_meeting(update, notes, hint)
            return

        if parts:
            await update.message.reply_text("\n".join(parts), parse_mode="Markdown")
        else:
            label = add_task_to_github(text)
            await update.message.reply_text(f"[{label}] `{text}`", parse_mode="Markdown")
    except Exception as e:
        try:
            label = add_task_to_github(text)
            await update.message.reply_text(f"[{label}] `{text}`", parse_mode="Markdown")
        except Exception as e2:
            await update.message.reply_text(f"Ошибка: {e2}")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    task_text = update.message.text.strip()
    if not task_text:
        return
    await route_text(update, task_text, context)


async def route_text(update: Update, task_text: str, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Route free-form text through all handlers (used by both text and voice)."""

    lower = task_text.lower()

    # Check for "инсайт:" prefix
    insight_match = re.match(r"^(?:инсайт|insight|💡)[:\s]+(.+)", lower, re.IGNORECASE)
    if insight_match:
        insight_text = task_text[insight_match.start(1):]  # preserve original case
        try:
            # Detect project
            project_id = detect_project(insight_text)
            if not project_id:
                project_id = detect_project_gpt(insight_text)
            if project_id == "unknown":
                project_id = "custody"  # default

            tags = detect_insight_tags(insight_text)
            save_insight(project_id, insight_text, tags)
            tags_str = " ".join(f"#{t}" for t in tags)
            await update.message.reply_text(
                f"💡 Инсайт сохранён в {PROJECTS[project_id]['name']}\n"
                f"{tags_str}\n\n{insight_text}",
            )
        except Exception as e:
            await update.message.reply_text(f"Ошибка сохранения инсайта: {e}")
        return

    # Check for "ссылка:" prefix
    link_match = re.match(r"^(?:ссылка|ссыл|link|🔗)[:\s]+(.+)", lower, re.IGNORECASE | re.DOTALL)
    if link_match:
        link_text = task_text[link_match.start(1):]
        try:
            # Extract URL(s) and description
            urls = re.findall(r"https?://\S+", link_text)
            if not urls:
                await update.message.reply_text("Не нашёл ссылку. Формат: `ссылка: https://... описание`", parse_mode="Markdown")
                return
            # Everything that's not a URL is the description
            description = re.sub(r"https?://\S+", "", link_text).strip(" ,:-—\n")

            # Detect project
            project_id = detect_project(link_text)
            if not project_id:
                project_id = detect_project_gpt(link_text + " " + description)
            if project_id == "unknown":
                project_id = "custody"

            tags = detect_insight_tags(description or urls[0])

            for url in urls:
                save_link(project_id, url, description or url, tags)

            tags_str = " ".join(f"#{t}" for t in tags)
            await update.message.reply_text(
                f"🔗 Сохранено в {PROJECTS[project_id]['name']}\n"
                f"{tags_str}\n\n" + "\n".join(urls) +
                (f"\n{description}" if description else ""),
            )
        except Exception as e:
            await update.message.reply_text(f"Ошибка сохранения ссылки: {e}")
        return

    # Check for "встреча:" prefix — Meeting notes
    meeting_match = re.match(r"^(?:встреча|meeting|митинг)[:\s]+(.+)", lower, re.IGNORECASE | re.DOTALL)
    if meeting_match:
        meeting_text = task_text[meeting_match.start(1):]
        # First line is project hint, but full text goes as notes
        first_line = meeting_text.strip().split("\n", 1)[0].strip()
        await process_meeting(update, meeting_text.strip(), first_line)
        return

    # Check for "контакт:/crm:" prefix — CRM commands
    crm_match = re.match(r"^(?:контакт|crm|црм)[:\s]+(.+)", lower, re.IGNORECASE | re.DOTALL)
    if crm_match:
        crm_text = task_text[crm_match.start(1):]
        try:
            data = gpt_parse_crm_command(crm_text)
            # Auto-detect project for CRM routing
            project_id = detect_project(crm_text)
            crm_target = get_crm_file(project_id) if project_id else None
            if data.get("action") == "update":
                name = update_crm_contact(data.get("name", ""), data, crm_target)
                if name:
                    changes = []
                    if data.get("next_action"):
                        changes.append(f"→ {data['next_action']}")
                    if data.get("ping_date"):
                        changes.append(f"📅 {data['ping_date']}")
                    if data.get("comment"):
                        changes.append(f"💬 {data['comment']}")
                    await update.message.reply_text(
                        f"✏️ Обновлён: *{name}*\n" + "\n".join(changes),
                        parse_mode="Markdown",
                    )
                else:
                    await update.message.reply_text("Не нашёл контакт для обновления. Попробуй точнее указать имя.")
            else:
                name = add_crm_contact(data, crm_target)
                await update.message.reply_text(
                    f"➕ Добавлен в CRM: *{name}*"
                    + (f" ({data.get('company', '')})" if data.get('company') else ""),
                    parse_mode="Markdown",
                )
        except Exception as e:
            await update.message.reply_text(f"Ошибка CRM: {e}")
        return

    # Check for "день:/итоги:" prefix — Evening review
    evening_match = re.match(r"^(?:день|итоги|вечер|review)[:\s]+(.+)", lower, re.IGNORECASE | re.DOTALL)
    if evening_match:
        evening_text = task_text[evening_match.start(1):]
        await process_evening(update, evening_text)
        return

    # Check for "перенеси" — task move command
    move_match = re.match(r"^(?:перенеси|перенести|передвинь|сдвинь|move)\s+(.+)", lower, re.IGNORECASE | re.DOTALL)
    if move_match:
        move_text = task_text[move_match.start(1):]
        try:
            data = gpt_parse_task_move(move_text)
            result = move_task(data["task_query"], data["new_date"])
            if result:
                await update.message.reply_text(f"📅 Перенесено: {result}")
            else:
                await update.message.reply_text("Не нашёл задачу. Попробуй точнее.")
        except Exception as e:
            await update.message.reply_text(f"Ошибка переноса: {e}")
        return

    # Check for "done/готово" prefix
    done_match = re.match(r"^(?:готово|done|✅|сделано)[:\s]+(.+)", lower)
    if done_match:
        query = done_match.group(1).strip()
        tasks = collect_open_tasks()
        # Fuzzy match
        matched = [t for t in tasks if query in t["text"].lower()]
        if len(matched) == 1:
            ok = toggle_task_in_file(matched[0]["file"], matched[0]["line_idx"], close=True)
            if ok:
                await update.message.reply_text(f"✅ Закрыто:\n~{matched[0]['text']}~", parse_mode="Markdown")
                return
        elif len(matched) > 1:
            # Show buttons to pick
            context.user_data["tasks"] = matched
            keyboard = []
            for i, t in enumerate(matched[:10]):
                keyboard.append([InlineKeyboardButton(t["text"][:45], callback_data=f"done:{i}")])
            await update.message.reply_text("Несколько совпадений, выбери:", reply_markup=InlineKeyboardMarkup(keyboard))
            return
        else:
            await update.message.reply_text("Не нашёл такую задачу. Попробуй /done для списка.")
            return

    await smart_process(update, task_text, context)


async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    try:
        voice_file = await update.message.voice.get_file()
        task_text = await transcribe_voice(voice_file)

        # Show transcription
        await update.message.reply_text(f"🎤 _{task_text}_", parse_mode="Markdown")

        # Route through unified processing (CRM, move, insights, tasks)
        await route_text(update, task_text, context)
    except Exception as e:
        await update.message.reply_text(f"Ошибка распознавания: {e}")


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return

    doc = update.message.document
    caption = update.message.caption or ""
    filename = doc.file_name or "document"

    await update.message.reply_text(f"📎 Получил `{filename}`, обрабатываю...", parse_mode="Markdown")

    try:
        # Download file
        tg_file = await doc.get_file()
        buf = bytearray()
        await tg_file.download_as_bytearray(buf)
        raw = bytes(buf)

        # Extract text based on file type
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if ext == "pdf":
            text = extract_pdf_text(raw)
        elif ext in ("txt", "md", "csv", "json", "py", "js", "ts", "sol"):
            text = raw.decode("utf-8", errors="replace")
        elif ext in ("xlsx", "xls"):
            text = extract_xlsx_text(raw)
        elif ext == "docx":
            text = extract_docx_text(raw)
        elif ext == "doc":
            text = raw.decode("utf-8", errors="replace")[:500] + "\n[Binary .doc — partial extraction]"
        else:
            text = raw.decode("utf-8", errors="replace")[:2000]

        if not text.strip():
            await update.message.reply_text("Не удалось извлечь текст из файла.")
            return

        # Detect project
        project_id = detect_project(caption) if caption else None
        if not project_id:
            project_id = detect_project(filename)
        if not project_id:
            project_id = detect_project_gpt(text, filename)

        if project_id == "unknown":
            await update.message.reply_text(
                "Не могу определить проект. Добавь подпись к файлу:\n"
                "`custody`, `sber`, `reksoft` или `blind-bets`",
                parse_mode="Markdown",
            )
            return

        # Summarize
        summary = summarize_document(text, filename)

        # Save to KB (with Telegram file_id for re-sending)
        save_to_kb(project_id, filename, text, summary, file_id=doc.file_id)

        project_name = PROJECTS[project_id]["name"]
        clean_summary = summary[:500].replace("*", "").replace("`", "").replace("_", "")
        await update.message.reply_text(
            f"📚 Сохранено в KB {project_name}\n\n"
            f"{filename}\n{clean_summary}",
        )

    except Exception as e:
        await update.message.reply_text(f"Ошибка обработки файла: {e}")


def extract_pdf_text(raw: bytes) -> str:
    import subprocess
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(raw)
        f.flush()
        result = subprocess.run(
            ["pdftotext", f.name, "-"],
            capture_output=True, text=True, timeout=30,
        )
    return result.stdout if result.returncode == 0 else ""


def extract_docx_text(raw: bytes) -> str:
    import zipfile
    import io
    import xml.etree.ElementTree as ET
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            xml_content = z.read("word/document.xml")
        tree = ET.fromstring(xml_content)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paragraphs = []
        for p in tree.iter(f"{{{ns['w']}}}p"):
            texts = [t.text for t in p.iter(f"{{{ns['w']}}}t") if t.text]
            if texts:
                paragraphs.append("".join(texts))
        return "\n\n".join(paragraphs)
    except Exception:
        return ""


def expand_search_query(query: str) -> list[str]:
    """Use GPT to expand search query with synonyms and abbreviations."""
    prompt = f"""Пользователь ищет «{query}» по базе знаний о крипто-кастоди бизнесе в России.
Расширь запрос: добавь синонимы, сокращения, альтернативные написания.
Например: «центробанк» → центробанк, ЦБ, Центральный банк, Банк России
«Тинькофф» → Тинькофф, Т-Банк, Tinkoff
«фандрайзинг» → фандрайзинг, fundraising, привлечение инвестиций, раунд

Ответь ТОЛЬКО списком через запятую, без объяснений. Включи оригинальный запрос."""

    try:
        answer = claude_complete(prompt, max_tokens=100, temperature=0)
        terms = [t.strip().lower() for t in answer.split(",") if t.strip()]
        return terms if terms else [query.lower()]
    except Exception:
        return [query.lower()]


def github_search_files(query: str) -> list[dict]:
    """Search across all files in the repo for a query string. Returns matches with context."""
    results = []
    search_terms = expand_search_query(query)

    def line_matches(line: str) -> bool:
        line_lower = line.lower()
        return any(term in line_lower for term in search_terms)

    # Search in main files
    for file_id in ["daily.md", "weekly.md", "backlog.md", "crm.md", "tracks.md"]:
        content_result = github_get_file(file_id)
        if not content_result:
            continue
        content = content_result[0]
        lines = content.split("\n")
        matches = []
        for i, line in enumerate(lines):
            if line_matches(line):
                start = max(0, i - 2)
                end = min(len(lines), i + 3)
                context_block = "\n".join(lines[start:end])
                matches.append(context_block)
        if matches:
            results.append({"file": file_id, "matches": matches[:3]})

    # Search in KB projects
    for pid, project in PROJECTS.items():
        index_path = f"{project['path']}/index.md"
        content_result = github_get_file(index_path)
        if not content_result:
            continue
        content = content_result[0]

        # Search index
        lines = content.split("\n")
        matches = []
        for i, line in enumerate(lines):
            if line_matches(line):
                start = max(0, i - 1)
                end = min(len(lines), i + 2)
                matches.append("\n".join(lines[start:end]))
        if matches:
            results.append({"file": f"KB/{project['name']}", "matches": matches[:3]})

        # Search docs inside project
        doc_links = re.findall(r"\[([^\]]+)\]\(([^\)]+\.md)\)", content)
        for doc_title, doc_filename in doc_links:
            doc_path = f"{project['path']}/{doc_filename}"
            doc_result = github_get_file(doc_path)
            if not doc_result:
                continue
            doc_content = doc_result[0]
            doc_lines = doc_content.split("\n")
            doc_matches = []
            for i, line in enumerate(doc_lines):
                if line_matches(line):
                    start = max(0, i - 2)
                    end = min(len(doc_lines), i + 3)
                    doc_matches.append("\n".join(doc_lines[start:end]))
            if doc_matches:
                results.append({"file": f"📎 {doc_title}", "matches": doc_matches[:3]})

    return results


def rank_search_results(results: list[dict]) -> list[dict]:
    """Rank results by number of matches, keep top 5."""
    ranked = sorted(results, key=lambda r: len(r["matches"]), reverse=True)
    return ranked[:5]


def clean_source_name(name: str) -> str:
    """Clean up file names for display."""
    # Remove KB/ prefix emoji
    name = re.sub(r"^📎\s*", "", name)
    # Remove timestamps and underscores from KB doc names
    name = re.sub(r"^\d{4}-\d{2}-\d{2}_\d{2}-\d{2}_", "", name)
    name = name.replace("_", " ").replace("  ", " ").strip()
    # Remove file extensions
    name = re.sub(r"\.(md|pdf|xlsx|docx)$", "", name)
    return name


def summarize_search_results(query: str, results: list[dict]) -> str:
    """Use GPT to create a smart summary of search results."""
    top = rank_search_results(results)

    raw_parts = []
    for r in top:
        source = clean_source_name(r["file"])
        raw_parts.append(f"=== {source} ===")
        # Deduplicate similar matches
        seen = set()
        for m in r["matches"][:2]:
            short = m[:100]
            if short not in seen:
                seen.add(short)
                raw_parts.append(m)
    raw_text = "\n\n".join(raw_parts)

    prompt = f"""Пользователь искал «{query}» по своей базе знаний (задачи, CRM, документы проектов).
Вот ТОП-5 самых релевантных фрагментов с контекстом:

{raw_text[:3000]}

Задача: объясни что КОНКРЕТНО говорится про «{query}» в этих документах.

ПРАВИЛА:
- Пиши ТОЛЬКО факты из фрагментов выше. НЕ додумывай.
- Каждый факт — с указанием источника в скобках
- Группируй по смыслу: что это, какая роль, какие действия
- Если есть задачи/дедлайны — выдели отдельно
- Если информации мало — скажи прямо
- Формат: короткие пункты, максимум 10 строк"""

    return claude_complete(prompt, max_tokens=400, temperature=0.1)


async def cmd_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id != ALLOWED_USER_ID:
        return
    if not context.args:
        await update.message.reply_text("Формат: `/search запрос`", parse_mode="Markdown")
        return

    query = " ".join(context.args)
    await update.message.reply_text(f"🔍 Ищу: _{query}_...", parse_mode="Markdown")

    try:
        results = github_search_files(query)
        if not results:
            await update.message.reply_text(f"Ничего не найдено по запросу «{query}»")
            return

        # GPT smart summary from top results
        summary = summarize_search_results(query, results)

        # Clean source list — top 5 unique
        top = rank_search_results(results)
        sources = [clean_source_name(r["file"]) for r in top]
        sources_str = "\n".join(f"  • {s}" for s in sources)

        total = sum(len(r["matches"]) for r in results)

        clean_summary = summary.replace("*", "").replace("`", "").replace("_", "")
        text = (
            f"🔍 *«{query}»* — {total} совпадений в {len(results)} файлах\n\n"
            f"{clean_summary}\n\n"
            f"📄 *Источники:*\n{sources_str}"
        )
        await update.message.reply_text(text[:4000], parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"Ошибка поиска: {e}")


def evening_analyze(text: str) -> dict:
    """Use Claude to analyze evening review text."""
    today = datetime.now()
    date_str = today.strftime("%Y-%m-%d")
    weekday = DAY_NAMES_SHORT[today.weekday()]

    # Get open tasks for matching
    open_tasks = []
    try:
        for fid in ["daily.md", "backlog.md", "weekly.md"]:
            result = github_get_file(fid)
            if result:
                for line in result[0].split("\n"):
                    s = line.strip()
                    if s.startswith("- [ ]") or re.match(r"^\d+\.\s*\[ \]", s):
                        t = re.sub(r"^(?:-|\d+\.)\s*\[ \]\s*", "", s)
                        if t and t != "[Task]":
                            open_tasks.append(t)
    except Exception:
        pass

    # Get habits from daily.md
    habits = []
    try:
        result = github_get_file("daily.md")
        if result:
            in_habits = False
            for line in result[0].split("\n"):
                if "## Habits" in line or "## Привычки" in line:
                    in_habits = True
                    continue
                if in_habits and line.strip().startswith("- [ ]"):
                    habit = re.sub(r"^-\s*\[ \]\s*", "", line.strip())
                    habits.append(habit)
                elif in_habits and line.strip().startswith("---"):
                    break
    except Exception:
        pass

    tasks_list = "\n".join(f"- {t}" for t in open_tasks[:30])
    habits_list = ", ".join(habits) if habits else "нет данных"

    prompt = f"""Проанализируй вечерний обзор дня от CEO.

Сегодня: {date_str} ({weekday})
Открытые задачи:
{tasks_list}

Привычки для отслеживания: {habits_list}

Сообщение пользователя:
{text}

Верни JSON (без markdown):
{{
  "done": [{{"query": "ключевые слова из ОТКРЫТЫХ задач для закрытия"}}],
  "crm_updates": [{{"name": "Имя", "company": "пусто", "action": "update", "next_action": "пусто", "comment": "что произошло", "ping_date": "пусто"}}],
  "habits": {{"done": ["название привычки которую выполнил"], "skipped": ["которую пропустил"]}},
  "summary": {{
    "done_text": "что сделано (кратко, 2-3 пункта)",
    "not_done_text": "что не сделано и почему",
    "mood": "настроение/энергия (1 слово)",
    "gratitude": "за что благодарен (из контекста)"
  }},
  "new_tasks": [{{"text": "новая задача если появилась", "priority": "normal", "date": "YYYY-MM-DD или пусто"}}]
}}

Правила:
- done.query: ТОЛЬКО ключевые слова из открытых задач выше. Не выдумывай.
- Если сказал "отправил Шмакову" — ищи задачу со словом Шмаков в открытых
- Если сказал "не успел Никиту" — НЕ закрывай, но упомяни в not_done_text
- habits: сопоставь с привычками по смыслу (если "покурил кальян" → Без кальяна = skipped)
- Пустые массивы НЕ включай"""

    answer = claude_complete(prompt, max_tokens=800, temperature=0.1)
    answer = re.sub(r"^```(?:json)?\s*", "", answer)
    answer = re.sub(r"\s*```$", "", answer)
    return json.loads(answer)


def save_journal_entry(date_str: str, data: dict) -> None:
    """Save evening journal entry to journal/YYYY-MM-DD.md."""
    summary = data.get("summary", {})
    content = (
        f"# {date_str}\n\n"
        f"## Что сделано\n{summary.get('done_text', '-')}\n\n"
        f"## Что не закончено\n{summary.get('not_done_text', '-')}\n\n"
        f"## Рефлексия\n"
        f"- *Состояние:* {summary.get('mood', '-')}\n"
        f"- *Хорошее:* {summary.get('gratitude', '-')}\n"
    )
    github_put_file(f"journal/{date_str}.md", content, f"journal: {date_str}")


def update_habit_tracker(habits_data: dict) -> None:
    """Update habit tracker in weekly.md."""
    result = github_get_file("weekly.md")
    if not result:
        return
    content, sha = result
    today_col = datetime.now().weekday()  # 0=Mon
    col_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    col_name = col_names[today_col]

    done_habits = [h.lower() for h in habits_data.get("done", [])]
    skipped_habits = [h.lower() for h in habits_data.get("skipped", [])]

    lines = content.split("\n")
    in_table = False
    header_cols = []
    col_idx = -1

    for i, line in enumerate(lines):
        if "| Habit |" in line:
            header_cols = [c.strip() for c in line.split("|")]
            try:
                col_idx = header_cols.index(col_name)
            except ValueError:
                return
            in_table = True
            continue
        if in_table and line.strip().startswith("|---"):
            continue
        if in_table and line.strip().startswith("|"):
            cells = line.split("|")
            if len(cells) > col_idx:
                habit_name = cells[1].strip().lower()
                mark = ""
                for dh in done_habits:
                    if dh in habit_name or habit_name in dh:
                        mark = " x "
                        break
                if not mark:
                    for sh in skipped_habits:
                        if sh in habit_name or habit_name in sh:
                            mark = " - "
                            break
                if mark:
                    cells[col_idx] = mark
                    lines[i] = "|".join(cells)
        elif in_table:
            break

    github_put_file("weekly.md", "\n".join(lines), f"habits: {datetime.now().strftime('%Y-%m-%d')}", sha)


def rotate_daily(top3: list[str] = None) -> None:
    """Create new daily.md with today's date and carry-over from yesterday."""
    today = datetime.now()
    date_str = today.strftime("%Y-%m-%d")
    weekday_en = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"][today.weekday()]

    # Read yesterday's daily for carry-over
    carry_over = []
    habits_section = ""
    result = github_get_file("daily.md")
    if result:
        content = result[0]
        # Check if already rotated today
        if date_str in content.split("\n")[0]:
            return
        for line in content.split("\n"):
            s = line.strip()
            if s.startswith("- [ ]") or re.match(r"^\d+\.\s*\[ \]", s):
                task = re.sub(r"^(?:-|\d+\.)\s*\[ \]\s*", "", s)
                if task and task != "[Task]":
                    carry_over.append(task)
        # Extract habits section
        in_habits = False
        habit_lines = []
        for line in content.split("\n"):
            if "## Habits" in line:
                in_habits = True
                habit_lines.append(line)
                continue
            if in_habits:
                if line.strip().startswith("---"):
                    break
                # Reset checkboxes
                habit_lines.append(line.replace("[x]", "[ ]"))
        if habit_lines:
            habits_section = "\n".join(habit_lines)

    # Build top-3
    top3_lines = ""
    if top3:
        for i, t in enumerate(top3[:3], 1):
            top3_lines += f"{i}. [ ] {t}\n"
    else:
        top3_lines = "1. [ ] \n2. [ ] \n3. [ ] \n"

    # Build carry-over section
    carry_section = ""
    if carry_over:
        carry_section = "\n## Carry-over\n" + "\n".join(f"- [ ] {t}" for t in carry_over) + "\n"

    if not habits_section:
        habits_section = ("## Habits\n"
            "- [ ] Без кальяна\n"
            "- [ ] Спорт/движение\n"
            "- [ ] Сон до 00:00\n"
            "- [ ] Контент-диета\n"
            "- [ ] Питание (норм еда)\n"
            "- [ ] Вода (2л)")

    new_daily = (
        f"# Today — {date_str} ({weekday_en})\n\n"
        f"## Top-3 (если сделаю только это — день успешный)\n"
        f"{top3_lines}"
        f"{carry_section}\n"
        f"{habits_section}\n\n"
        f"---\n"
        f"*Last updated: {date_str}*\n"
    )

    sha = result[1] if result else None
    github_put_file("daily.md", new_daily, f"daily: {date_str}", sha)


def collect_morning_context() -> dict:
    """Gather all data needed for morning briefing."""
    today = datetime.now().date()
    today_str = today.strftime("%d.%m")
    weekday = DAY_NAMES_SHORT[today.weekday()]
    ctx = {"date": str(today), "weekday": weekday}

    # Yesterday's daily — unfinished tasks
    daily_result = github_get_file("daily.md")
    if daily_result:
        yesterday_open = []
        yesterday_done = []
        for line in daily_result[0].split("\n"):
            s = line.strip()
            if s.startswith("- [ ]") or re.match(r"^\d+\.\s*\[ \]", s):
                task = re.sub(r"^(?:-|\d+\.)\s*\[ \]\s*", "", s)
                if task and task != "[Task]":
                    yesterday_open.append(task)
            elif s.startswith("- [x]") or re.match(r"^\d+\.\s*\[x\]", s):
                task = re.sub(r"^(?:-|\d+\.)\s*\[x\]\s*", "", s)
                if task:
                    yesterday_done.append(task)
        ctx["yesterday_open"] = yesterday_open
        ctx["yesterday_done"] = yesterday_done

    # Backlog — today's tasks, overdue, upcoming
    backlog_result = github_get_file(BACKLOG_FILE)
    if backlog_result:
        today_tasks = []
        overdue = []
        upcoming = []
        no_date = []
        for line in backlog_result[0].split("\n"):
            s = line.strip()
            if "[ ]" not in s:
                continue
            task = re.sub(r"^(?:-|\d+\.)\s*\[ \]\s*", "", s)
            date_m = re.search(r"\(→\s*[^\d]*(\d{2}\.\d{2})\)", s)
            if date_m:
                task_clean = re.sub(r"\s*\(→[^)]+\)", "", task)
                try:
                    task_date = datetime.strptime(f"{date_m.group(1)}.{today.year}", "%d.%m.%Y").date()
                    if task_date == today:
                        today_tasks.append(task_clean)
                    elif task_date < today:
                        days = (today - task_date).days
                        overdue.append(f"{task_clean} ({days}дн просрочено)")
                    elif (task_date - today).days <= 3:
                        upcoming.append(f"{task_clean} ({date_m.group(1)})")
                except ValueError:
                    pass
            else:
                no_date.append(task)
        ctx["today_tasks"] = today_tasks
        ctx["overdue_tasks"] = overdue
        ctx["upcoming_tasks"] = upcoming
        ctx["no_date_tasks"] = no_date[:5]

    # CRM pings — across all CRM files
    crm_today = []
    crm_overdue = []
    seen_crm = set()
    for proj in PROJECTS.values():
        crm_file = proj.get("crm", CRM_FILE)
        if crm_file in seen_crm:
            continue
        seen_crm.add(crm_file)
        crm_result = github_get_file(crm_file)
        if crm_result:
            contacts = parse_crm(crm_result[0])
            for c in contacts:
                if c["ping_date"]:
                    if c["ping_date"] == today:
                        crm_today.append(f"{c['name']} ({c['company']}) → {c['next_action']}")
                    elif c["ping_date"] < today:
                        days = (today - c["ping_date"]).days
                        crm_overdue.append(f"{c['name']} ({c['company']}) → {c['next_action']} ({days}дн)")
    ctx["crm_today"] = crm_today
    ctx["crm_overdue"] = crm_overdue

    # Weekly focus
    weekly_result = github_get_file("weekly.md")
    if weekly_result:
        for line in weekly_result[0].split("\n"):
            if line.strip().startswith(">"):
                ctx["weekly_focus"] = line.strip().lstrip("> ")
                break

    return ctx


def generate_morning_briefing(ctx: dict) -> str:
    """Use Claude to generate smart morning briefing."""
    prompt = f"""Ты — персональный ассистент CEO крипто-стартапа. Сгенерируй утренний брифинг.

Сегодня: {ctx['date']} ({ctx['weekday']})
Фокус недели: {ctx.get('weekly_focus', 'не задан')}

Вчера сделано: {', '.join(ctx.get('yesterday_done', [])) or 'нет данных'}
Вчера НЕ сделано (carry-over): {', '.join(ctx.get('yesterday_open', [])) or 'всё сделано'}

Задачи на сегодня (backlog): {', '.join(ctx.get('today_tasks', [])) or 'нет'}
Просроченные задачи: {', '.join(ctx.get('overdue_tasks', [])) or 'нет'}
Ближайшие (2-3 дня): {', '.join(ctx.get('upcoming_tasks', [])) or 'нет'}
Без даты: {', '.join(ctx.get('no_date_tasks', [])) or 'нет'}

CRM пинги сегодня: {', '.join(ctx.get('crm_today', [])) or 'нет'}
CRM просроченные: {', '.join(ctx.get('crm_overdue', [])) or 'нет'}

Сгенерируй КОРОТКИЙ брифинг для Telegram (макс 15 строк):

1. TOP-3 приоритета на сегодня (самое важное из всех источников)
2. Carry-over: незакрытые вчерашние задачи (если >3 дней — пометь)
3. CRM: кому написать/позвонить сегодня
4. Quick wins: что можно сделать за 5 мин прямо сейчас

Формат: чистый текст для Telegram Markdown. Без ```блоков```. Будь конкретным, не общие фразы. Используй * для bold."""

    return claude_complete(prompt, max_tokens=600, temperature=0.3)


async def morning_reminder(bot) -> None:
    """Send smart morning briefing at 9:00 MSK + rotate daily."""
    chat_id = ALLOWED_USER_ID
    try:
        ctx = collect_morning_context()
        briefing = generate_morning_briefing(ctx)

        # Rotate daily.md with top-3 from context
        top3 = ctx.get("today_tasks", [])[:3]
        if not top3:
            top3 = ctx.get("overdue_tasks", [])[:3]
        # Clean up task text for top-3
        top3_clean = [re.sub(r"\s*\(\d+дн.*?\)", "", t) for t in top3]
        rotate_daily(top3_clean if top3_clean else None)

        await bot.send_message(
            chat_id=chat_id,
            text=briefing[:4000],
            parse_mode="Markdown",
        )
    except Exception as e:
        print(f"Morning reminder error: {e}", flush=True)
        try:
            await bot.send_message(
                chat_id=chat_id,
                text=f"Доброе утро! Ошибка генерации брифинга: {e}\nИспользуй /today и /backlog",
            )
        except Exception:
            pass


def extract_xlsx_text(raw: bytes) -> str:
    try:
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                vals = [str(c) if c is not None else "" for c in row]
                if any(vals):
                    lines.append("\t".join(vals))
        return "\n".join(lines)
    except Exception:
        return ""


def main() -> None:
    print("Clearing webhook and waiting for old instance...", flush=True)
    try:
        httpx.post(
            f"https://api.telegram.org/bot{TOKEN}/deleteWebhook",
            json={"drop_pending_updates": True},
            timeout=10,
        )
    except Exception as e:
        print(f"Webhook cleanup failed: {e}", flush=True)
    time.sleep(3)

    from datetime import time as dt_time, timezone, timedelta

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("today", cmd_today))
    app.add_handler(CommandHandler("week", cmd_week))
    app.add_handler(CommandHandler("backlog", cmd_backlog))
    app.add_handler(CommandHandler("crm", cmd_crm))
    app.add_handler(CommandHandler("meetings", cmd_meetings))
    app.add_handler(CommandHandler("morning", cmd_morning))
    app.add_handler(CommandHandler("evening", cmd_evening))
    app.add_handler(CommandHandler("tracks", cmd_tracks))
    app.add_handler(CommandHandler("insights", cmd_insights))
    app.add_handler(CommandHandler("project", cmd_project))
    app.add_handler(CommandHandler("done", cmd_done))
    app.add_handler(CommandHandler("search", cmd_search))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    # Schedule morning reminder at 9:00 Moscow time (UTC+3) via asyncio
    import asyncio
    moscow_tz = timezone(timedelta(hours=3))

    async def reminder_loop(application):
        """Background loop that sends morning reminder at 9:00 MSK daily."""
        while True:
            now = datetime.now(moscow_tz)
            target = now.replace(hour=9, minute=0, second=0, microsecond=0)
            if now >= target:
                target = target + timedelta(days=1)
            wait_seconds = (target - now).total_seconds()
            print(f"Next morning reminder in {wait_seconds/3600:.1f}h", flush=True)
            await asyncio.sleep(wait_seconds)
            try:
                await morning_reminder(application.bot)
            except Exception as e:
                print(f"Morning reminder failed: {e}", flush=True)

    async def post_init(application):
        asyncio.create_task(reminder_loop(application))

    app.post_init = post_init
    print("Morning reminder scheduled for 9:00 MSK (asyncio)", flush=True)

    print(f"Bot started, repo: {GITHUB_REPO}", flush=True)
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
